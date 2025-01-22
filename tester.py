
import pandas as pd
import os


filpathdef = 'my_data/001.xlsx'
file_path = 'my_data/my_file.txt'
filenamedef = '1'
fileexepdef ='xlsx'

def squararia(w,l):
    return w*0.001*l*0.001;

def volumeofarea(a,l):
    return a*l*0.001

def circarea(d):
    d = float(str(d))*0.001
    a = 3.14*d*d
    return a

def cylindrevol(od,tickness):
    return volumeofarea(circarea(od),tickness)



def massofvolumeofmaterial(vol,den):
    m = vol*den*1000
    return m

def calcpricebymass(mass,ppm):
    return mass*ppm

def structmakinerfromexcel(sn):
    dbfilename='mainDB_.xlsx'
    hsheet = pd.read_excel(dbfilename, sheet_name =str(sn.lower()), index_col=None, header=0)
    df = pd.DataFrame(hsheet)    
    data=[]
    r=[]
    s=[]
    for row in df.itertuples():
        r = row[1:]  # Skip the index
        data.append(r)
    return data    

#print(str(structmakinerfromexcel('material')))

def findmaterialifpredefined(mat='s316'):
    data = structmakinerfromexcel('material')
    m = str(mat).lower()
    if len(m)>3:
        m=m[0:3]
        
    
    for d in data:
        s=str(d[2]).lower()
        if len(s)>2:
            s=s[0:3]
        
        if s==m:
            return d
    ndd=[]
    ndd.append(0)
    ndd.append(str(mat))
    ndd.append(mat)
    ndd.append(mat)
    ndd.append(5)
    ndd.append(5000000)
    return ndd

def getmaterialnamefa(mat='s316'):
    mf = findmaterialifpredefined(mat)
    return str(mf[3])

def getmaterialnamepriceunit(mat='s316'):
    mf = findmaterialifpredefined(mat)
    return str(mf[5])
        
   
def getrupturelayers(rtype='reverse'):
    data = structmakinerfromexcel('reverse')
    dd=[]
    r=[]
    mr=[]
    m=[]
    for d in data:
        r=[]
        mr=[]
        for i in range(5):
            m=[]
            if not pd.isna(d[3+i*2]):
                m=[]
                m.append(d[3+i*2])
                m.append(d[3+1+i*2])
                mr.append(m)
        r.append(mr)        
        m=[]
        m.append(d[13])
        m.append(d[14])
        m.append(d[15])
        r.append(m)
        m=[]
        m.append(d[2])
        m.append(d[16])
        m.append(d[17])        
        r.append(m)
        dd.append(r)
    return dd 


def findsizeequalityvalue(rtype='reverse',rsize=2):
    data = getrupturelayers(rtype)
    sized=[]
    for test in data:
        tc = test[2]
        if tc[0]==rsize:
            sized.append(test)
    return sized
            
 
def sortingbyrburstpressure(rtype='reverse',rsize=2,rbp=5):
    datas= findsizeequalityvalue(rtype,rsize)
    data=[]
    for rds in datas:
        rd = rds[1]
        rds.append(abs(rbp-float(rd[0])))
        data.append(rds)
    sdata=     sorted(data,key=lambda x:x[3])
    return sdata
     
def findneardesignedruptureperv(rtype='reverse',rsize=2,rbp=5):
    data = sortingbyrburstpressure(rtype,rsize,rbp)
    if len(data)>0:
        return data[0]
    else:
        rsize=rsize+2
        data = sortingbyrburstpressure(rtype,rsize,rbp)
        return data[0]
 
def isitvalidsizeofelement(esize=4):
    data = structmakinerfromexcel('size')
    for d in data:
        if float(d[2])== float(esize):
            return True
    return False
    
    
        
def getdimensionbysizetype(element='rupture',etype='reverse',esize=2):
    data = structmakinerfromexcel('size')
    etype=etype.lower()
    etype=etype.lower()
    r=[]
    for d in data:
        if d[0]==element:
            
            if  etype=='' or d[1]== etype:
                if (esize == 0) or (d[2] == esize):
                    for i in range(4,len(d)):
                        if d[i]!=0:
                            r.append(float(d[i]))
                            
                
    return r
                
def calcrawsheetforruptures(rodl=118,qty=5,shw=2000,shh=1000):
    rod=rodl[0]
    lsm=10
    lcm =5
    xsepm=5
    rsq=1
    rsw = shw - lsm*2
    rsh= shh -2*lsm       
    rrod = rod + lcm*2
    rawsheetwide = (rrod+xsepm)*qty+2*lsm
    rawsheetheight = rrod +lsm*2
    rawdim = []
    rawdim.append(int(rawsheetwide))
    rawdim.append(int(rawsheetheight))
    rawdim.append(rsq)
     
    rprm = int(rsw/(rrod+xsepm))
    if qty < rprm:
        return rawdim  
    rawdim = []
    rawsheetwide=shw
    ystagerratio = pow(3,0.5)*0.5
    
    
    
    rprn = rprm-1
    col=0
    row=0
    for i in range(qty):
        col+=1
        if col >= (rprm-(row%2)):
            row+=1
            col = 0
    
    rawsheetheight +=  (rrod) *ystagerratio*row
    if rawsheetheight> shh:
        rsq = int(rawsheetheight/shh)
        rsq+=1
    rawsheetheightover = rawsheetheight % shh
    rawdim = []
    rawdim.append(rawsheetwide)
    rawdim.append(int(rawsheetheight))
    rawdim.append(rsq)
    rawdim.append(int(rawsheetheightover))
    return rawdim         
                
    
                    
def getreadmtoitemorop(element='rupture'):
    data = structmakinerfromexcel('mto')
    r=[]
    it=[]
    rd=[]
    for d in data:
        if d[0]==element:
            r=[]
            it=[]
            it.append(d[2])
            it.append(d[3])
            r.append(it)
            it=[]
            it.append(d[4])
            it.append(d[5])
            r.append(it)
            it=[]
            it.append(d[6])
            it.append(d[7])
            r.append(it)    
            rd.append(r)
    return rd   

def getreadmtoheader(l='Fa'):
    data = structmakinerfromexcel('mto')
    r=[]
    it=[]
    for d in data:
        if d[0]=='mtoheader':
            
            if l=='Fa':
            
                it.append(d[3])
            else:
                it.append(d[2])
            r.append(it)
            
    return it

def getdensityofmaterial(mat='s316'):
    mf = findmaterialifpredefined(mat)
    return str(mf[4])    

def getpriceofmaterialkg(mat):
    mf = findmaterialifpredefined(mat)
    return str(mf[5])  

def getunitpriceformtoitem(mtoitem,itemnum):
    unit = mtoitem[itemnum]
    return unit[2]
 
ruplayermaterialsdef=['s316','s316','ptfe']
ruptomtodef=['reverse',2,5,50,ruplayermaterialsdef,5,False,False,False,False,False,False,False]   
rawsheetwidemainlayers=2000
rawsheetwideseallayer=1200
rawsheetheightmainlayers=1000
rawsheetheightseallayer=4000
def makeandwritemtoforrupture(ruptomto=ruptomtodef):
    SHW=rawsheetwidemainlayers
    SHH =1000
    TSHW=1200
    rtype =ruptomto[0]
    rsize=ruptomto[1]
    rbp=ruptomto[2]
    rbt=ruptomto[3]
    lm1= ruptomto[4]
    qty= ruptomto[5]
    sensor=ruptomto[6]
    wirecut=ruptomto[7]
    analysmat=ruptomto[8]
    ship=ruptomto[9]
    box=ruptomto[10]
    tag=ruptomto[11]
    water_laser=ruptomto[12]
    mtoh = getreadmtoheader()
    mtoit = getreadmtoitemorop('rupture')
    data = findneardesignedruptureperv(rtype,rsize,rbp)
    ename = ''
    indexconter=0
    mtorows=[]
    mtro=[]
    eprppmat = ''
    if len(data)==0:
        return eprppmat
    laym=[]
    layers=data[0]
    
    seallayerfounded=False
    
    for l in layers:
        isitseallayer=False
        lm = l[0]
        if lm.find('pt')>-1 or lm.find('pv')>-1 or lm.find('pl')>-1 or lm.find('tef')>-1 or lm.find('fep')>-1 :
            isitseallayer=True 
            seallayerfounded=True
        lt = l[1]
        ename = ''
        en = mtoit[0]
        nn = en[0]
        if isitseallayer:
            en = mtoit[2]
            nn = en[0]            
            enameln = str(nn[1])
            SHW=rawsheetwideseallayer
            SHH = rawsheetheightseallayer
        else:
            en = mtoit[1]
            nn = en[0]            
            enameln = str(nn[1])
            SHW = rawsheetwidemainlayers
            SHH = rawsheetheightmainlayers
        ename = ename.__add__(enameln)    
        ename = ename.__add__(" ")
        ename = ename.__add__(str(getmaterialnamefa(lm)))
        
        ruprawdia = getdimensionbysizetype('rupture',rtype,rsize)
        rawsheetdim = calcrawsheetforruptures(ruprawdia,qty,SHW,SHH)        
        
        eprppmat=ename
        eprppmat=eprppmat.__add__(' , Thickness: ')
        eprppmat=eprppmat.__add__(str(lt))
        eprppmat=eprppmat.__add__(' mm  , Dim: ')
        eprppmat=eprppmat.__add__(str(rawsheetdim[0]))
        eprppmat=eprppmat.__add__(' * ')
        eprppmat=eprppmat.__add__(str(rawsheetdim[1]))
        eprppmat=eprppmat.__add__(' mm ')
        
        area = squararia(rawsheetdim[0],rawsheetdim[1])
        vol = volumeofarea(area,lt)
        dens = float(getdensityofmaterial(lm))
        
        
        msslayer = massofvolumeofmaterial(vol,dens)
         
        mtro=[]
        indexconter+=1
        mtro.append(indexconter)
        mtro.append(ename)
        mtro.append(eprppmat)
        mtro.append(1)
        unitpprice = en[1]#getunitpriceformtoitem(mtoit,1)
        mtro.append(unitpprice[0])
        mtro.append(msslayer)
        matpriceunit = float(getpriceofmaterialkg(lm))
        mtro.append(matpriceunit)
        matprice = calcpricebymass(matpriceunit,msslayer)
        mtro.append(matprice)
        
            
        mtorows.append(mtro)       
    
    if seallayerfounded==False:
        en = mtoit[2]
        nn = en[0]        
        enameln = str(nn[1])
        SHW=rawsheetwideseallayer
        SHH = rawsheetheightseallayer 
        ename=''
        ename = ename.__add__(enameln)    
        ename = ename.__add__(" ")
        ename = ename.__add__(str(getmaterialnamefa('teflon')))
        
        eprppmat=''
        eprppmat=ename
        eprppmat=eprppmat.__add__(' , Thickness: ')
        eprppmat=eprppmat.__add__(str(0.5))
        
        ruprawdia = getdimensionbysizetype('rupture',rtype,rsize)
        rawsheetdim = calcrawsheetforruptures(ruprawdia,qty,SHW,SHH)
        
        eprppmat=eprppmat.__add__(' mm  , Dim: ')
        eprppmat=eprppmat.__add__(str(rawsheetdim[0]))
        eprppmat=eprppmat.__add__(' * ')
        eprppmat=eprppmat.__add__(str(rawsheetdim[1]))
        eprppmat=eprppmat.__add__(' mm ')
        area = squararia(rawsheetdim[0],rawsheetdim[1])
        vol = volumeofarea(area,lt)
        dens = float(getdensityofmaterial(lm))
    
    
        msslayer = massofvolumeofmaterial(vol,dens)
    
        mtro=[]
        indexconter+=1
        mtro.append(indexconter)
        mtro.append(ename)
        mtro.append(eprppmat)
        mtro.append(1)
        unitpprice = en[1]#(mtoit,1)
        mtro.append(unitpprice[0])
        mtro.append(msslayer)
        matpriceunit = float(getpriceofmaterialkg(lm))
        mtro.append(matpriceunit)
        matprice = calcpricebymass(matpriceunit,msslayer)
        mtro.append(matprice)
    
    
        mtorows.append(mtro)               
        
    Q=1
    eprppmat=" "
    if analysmat==True:
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,3,1,indexconter))
        
    if water_laser==True:
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,5,1,indexconter))
   
    if sensor==True:
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,8,1,indexconter))
        
    if wirecut==True:
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,4,1,indexconter))
   
        
    if tag==True:
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,7,1,indexconter))
        
    if box==True:
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,11,1,indexconter))
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,12,1,indexconter))
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,13,1,indexconter))
       
    if ship==True:
        indexconter+=1
        mtorows.append(anotherselementofmto(mtoit,14,1,indexconter))
        
    mtorows.append(["","","","","","","Total Price",'=SUM(H2:H12)'])
    df1 = pd.DataFrame(mtorows,columns=mtoh)  
    mtoTitle = "Project ".__add__('Rupture Disks, type: ,').__add__(rtype).__add__(' , Size: ').__add__(str(rsize)).__add__(' inch ,  Burst Pressure: ').__add__(str(rbp)).__add__(' bar, Qty: ').__add__(str(qty))
    ffxn = getmyfilename(rt=rtype,rs=rsize,rb=rbp,rq=qty)
    return writedataframedstylishtofile(df1,ffxn,mtoTitle)
        
    return mtorows
    
def anotherselementofmto(mtoit,itnum=4,qty=1,i=4):
      
    ow=[]
    en = mtoit[itnum]
    nn = en[0]        
    enameln = str(nn[1])
    ow.append(i) 
    ow.append(enameln)      
    ow.append(" ")
    ow.append(qty)
    mm= en[1] 
    ow.append('-')
    
    ow.append("-")
    ow.append(mm[1])
    tp = float(mm[1])*qty
    ow.append(str(int(tp)))
    return ow 


def getmyfilename(pt='my_data/',fn='1',rt='reverse',rs=4,rb=11,rq=10):
    filename=""
    if not os.path.exists(pt): 
          
        # if the demo_folder directory is not present  
        # then create it. 
        os.makedirs(pt)     
    if fn.find('__')>-1:
        filename = fn[0:fn.find('__')]
    if fn.find('.')>-1:
        filename = fn[0:fn.find('.')]
        fileexo = fn[fn.find('.'):len(fn)]
    else:
        fileexo = fileexepdef
        for f in fn :
            if f in ('0123456789'):
                filename = filename.__add__(f) 
    
                
    file_path_name = pt.__add__(filename).__add__('__type ').__add__(str(rt)).__add__('_size ').__add__(str(rs)).__add__('_BP ').__add__(str(rb)).__add__('_QTY ').__add__(str(rq)).__add__('.').__add__(fileexo)
    # Check if the file exists
    while os.path.exists(file_path_name):
        filename = str(int(filename)+1)
        file_path_name = pt.__add__(filename).__add__('__type ').__add__(str(rt)).__add__('_size ').__add__(str(rs)).__add__('_BP ').__add__(str(rb)).__add__('_QTY ').__add__(str(rq)).__add__('.').__add__(fileexo)
    return file_path_name

         
def excelmtofilestylerfarsi(fn='mto',shn='_mto',title='Project'):
    
 
    from openpyxl import load_workbook 
    from openpyxl.styles import Font, PatternFill, Alignment ,Border , Side
    from openpyxl.utils.dataframe import dataframe_to_rows 
    # Sample DataFrame to write to Excel 
    # Write DataFrame to Excel file 
    file_path = fn
    
    # Load the workbook and the sheet 
    wb = load_workbook(file_path) 
    
    sheet = wb[shn] 
    sheet.sheet_properties.rightToLeft = True 
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))      
    borderH = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))       
    # Set font for the rest of the sheet and wrap text 
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=8): 
        for cell in row: cell.font = Font(name="B Nazanin", size=14) 
        cell.alignment = Alignment(wrap_text=True) 
        cell.border = border
        # Wrap text in each cell 
        # Set header row (Row 2) 
    header_row = sheet[1] 
    for cell in header_row: 
        cell.font = Font(size=16,bold=True) 
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
        cell.fill = PatternFill(start_color="90A4FA", end_color="90A4FA", fill_type="solid") 
        # Light gray background 
    # Set row background colors (odd/even rows) 
      
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8), start=2):
        for cell in row: 
            if i % 2 == 0: 
                cell.fill = PatternFill(start_color="9BEDEF", end_color="9BEDEF", fill_type="solid")
                cell.border = border
                # Even row light gray 
            else: 
                cell.fill = PatternFill(start_color="CBF298", end_color="CBF298", fill_type="solid")
                cell.border = border
                # Odd row white 
    # Set sheet direction to right to left 
    sheet.sheet_properties.rightToLeft = True 
    
    # Resize columns to fit the content 
    for col in sheet.columns:
        max_length = 0 
        column = col[0].column_letter 
        # Get the column name (e.g., 'A', 'B', etc.) 
        for cell in col: 
            try: 
                if len(str(cell.value)) > max_length: 
                    max_length = len(cell.value) 
            except: 
                pass 
            adjusted_width = (max_length + 2) 
            sheet.column_dimensions[column].width = adjusted_width 
   
            
    
   

    # Select the active sheet
    sheet = wb.active

    # Insert an empty row at the top
    sheet.insert_rows(1)

    
    # Set title in the first row and merge cells 
    sheet.merge_cells('A1:H1')
    title_cell = sheet['A1'] 
    title_cell.value = title
    title_cell.font = Font(name="Arial",size=18, bold=True) 
    title_cell.alignment = Alignment(horizontal='center', vertical='center') 
    title_cell.fill = PatternFill(start_color="FFFF25", end_color="FFFF25", fill_type="solid") 
    #title_cell.border = borderH
    # Yellow background
    print("Excel file styled and saved with wrapped text and auto-resized columns.") 
    
    
    # Save the workbook 
    wb.save(file_path)    

  
def writedataframedstylishtofile(df,fn,title):
    df.to_excel(fn,sheet_name='_mto',index=False) 
    excelmtofilestylerfarsi(fn,'_mto',title)    
    return fn

            

def drawrupturesinplate(gr,qty=8,rod=155,rph=1000,rawplatewide=2000):  
    gm = 5 #graphic margin that allstart at this
    gblx=gr.BottomLeft[0]
    gbly=gr.BottomLeft[1]
    gw=gr.Size[0]
    gh=gr.Size[1]
    gtrx=gblx+gw
    gtry=gbly+gh
    rpblx=gblx+gm
    rpbly=gbly+gm
    rptrym=gtry-gm
    rptrxm=gtrx-gm
    rpgw=rptrxm-rpblx
    rpgh=rptrym-rpbly
    xsep=5
    ysep=5
    rpsm=10
    lcm=2
    halfdia=rod/2
    rad=halfdia
    rpr = int((rawplatewide -25)/(rod+10))
    screenviewWiderawplate= rptrxm-rpblx
    ratio = screenviewWiderawplate/rawplatewide
    screenviewheightrawplate = int(rph*ratio)
    rptrym=rph*ratio+gm
    rpgh=rptrym
    drr=int(rad*ratio)
    rcrad=drr+lcm
    cpfdrx = rpblx+rpsm+lcm+drr
    cpfdry=  rpbly+rpsm+lcm+drr
    usedplateblx=rpblx
    usedplatebly = rpbly
    
    cpsepx = drr+lcm+xsep+lcm+drr
    cpsepy = int(ysep+(drr+lcm+lcm+drr)*pow(3,0.5)*0.5)
    cpfrsrx= int(cpfdrx+cpsepx*0.5)
    col = 0
    row = 0
    gr.draw_rectangle((rpblx,rpbly),(rpgw,rpgh), fill_color='blue', line_color='green', line_width=1)
    usedplatetrx =usedplateblx
    usedplatetry = cpfdry*2
    
    
    
    
    
    if qty>= rpr:
        usedplatetrx = rptrxm
        upgw=rpgw
    else:
        usedplatetrx = (cpfdrx*2)+ (qty-1)*(cpsepx-rpsm)
        upgw=(cpfdrx*2)+ (qty-1)*(cpsepx-rpsm)
    if upgw> rpgw:
        usedplatetrx = rptrxm
        upgw=rpgw
        
    rowtqty=1
    stagering=1
    if  (((rpr-1)* cpsepx )+ cpfdrx*2) < rpgw:
        stagering = 0
        rowtqty = 2*rpr
    else:
        rowtqty = 2*rpr-1
    rowt= int(qty/rowtqty)
    rowqtymod = qty-(rowt*rowtqty)
    rowi=0
    if rowqtymod>=rpr:
        rowi = 2
    elif rowqtymod>0:
        rowi=1
    else:
        rowi=0
    rowu = rowt*2+rowi
    upgh= (rowu-1)*cpsepy+(cpfdry-rpsm)*2
    upgw=rpgw    
    gr.draw_rectangle((usedplateblx,usedplatebly),(upgw,upgh), fill_color='magenta', line_color='red', line_width=0)
    cpfdrx0=cpfdrx
    for i in range(qty):
        col+=1
        if col == rpr-(row%2)*stagering:
            row+=1
            col = 1
            if row%2==1:
                cpfdrx0=cpfrsrx
            else:
                cpfdrx0=cpfdrx                
            
        x= ((col -1)* cpsepx )+ cpfdrx0
        if x+drr+lcm+gm > rptrxm:
            row+=1
            col = 1
            if row%2==1:
                cpfdrx0=cpfrsrx
            else:
                cpfdrx0=cpfdrx             
            
        x= ((col -1)* cpsepx )+ cpfdrx0    
            
        
        y=row*cpsepy+cpfdry
        gr.draw_circle(center_location=(x,y), radius=drr, fill_color='red', line_color='yellow', line_width=lcm) 
        
    
def makemtorowsforholder(htype='reverse',hsize=2,qty=2,material='s316'):
    htype=htype.lower()
    hdim=getdimensionbysizetype(element='holder',etype=htype,esize=hsize)
    hod=hdim[0]
    hid=hdim[1]
    hticktot=hdim[2]
    psth=dim[4]
    vsth=d[5]
    hthickness=hdim[3]
    
    vsroundedth = (int(vsth/10)+1)*10
    psroundedth = (int(psth/10)+1)*10
    hroundeddia = (int(hod/10)+1)*10
    
    harea = circarea(hod)
    hden = getdensityofmaterial(material)    
    vsshaftvol = volumeofarea(harea,vsroundedth)
    psshaftvol = volumeofarea(harea,psroundedth)
    vsmass = massofvolumeofmaterial(vsshaftvol,hden)
    psmass = massofvolumeofmaterial(psshaftvol,hden)
    matnamefa = getmaterialnamefa(material)
    matpriceperunit = getpriceofmaterialkg(material)
    
    mtoh = getreadmtoheader()
    mtoit = getreadmtoitemorop('holder')
    data = findneardesignedruptureperv(rtype,rsize,rbp)
    ename = ''
    indexconter=0
    mtorows=[]
    mtro=[]
    eprppmat = ''
    if len(data)==0:
        return eprppmat
    ename = ''
    en = mtoit[1]
    nn = en[0]
    enameln = str(nn[0])
    ename = ename.__add__(enameln)    
    
    
    
    
    eprppmat=ename
    eprppmat=eprppmat.__add__(' , OD: ')
    eprppmat=eprppmat.__add__(str(hod))
    eprppmat=eprppmat.__add__(' mm  , len: ')
    eprppmat=eprppmat.__add__(str(vsroundedth))
    eprppmat=eprppmat.__add__(' * ')
    eprppmat=eprppmat.__add__(str(rawsheetdim[1]))
    eprppmat=eprppmat.__add__(' mm ')
    
    area = squararia(rawsheetdim[0],rawsheetdim[1])
    vol = volumeofarea(area,lt)
    dens = float(getdensityofmaterial(lm))
    
    
    msslayer = massofvolumeofmaterial(vol,dens)
     
    mtro=[]
    indexconter+=1
    mtro.append(indexconter)
    mtro.append(ename)
    mtro.append(eprppmat)
    mtro.append(1)
    #unitpprice = en#getunitpriceformtoitem(mtoit,1)
    unitpprice = en[1]
    mtro.append(unitpprice[0])
    mtro.append(msslayer)
    matpriceunit = float(getpriceofmaterialkg(lm))
    mtro.append(matpriceunit)
    matprice = calcpricebymass(matpriceunit,msslayer)
    mtro.append(matprice)
    
        
    mtorows.append(mtro)           